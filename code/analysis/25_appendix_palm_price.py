"""Appendix figure: World Bank monthly palm-oil price, 2007--2015 (continuous).

The earlier draft chart had only the two IFLS-fielding sub-windows plotted
(2007:6--2008:9 and 2014:8--2015:12) with a six-year gap in between, because
the analysis only needs prices that fall inside the 3-month lag-window used
to construct the palm-shock variable. For the paper appendix we want a
continuous series so the reader sees the full price trajectory --- the 2008
commodity-boom peak and post-GFC crash, the 2010--2011 recovery, the
2012--2014 secular decline, and the 2014--2015 China-demand collapse.

The script:
  1. Attempts to download the latest WB Pink Sheet monthly prices from the
     official URL and extract the Palm oil column (Malaysia FOB, USD/MT).
     Cached locally as `palm_oil_monthly_2007_2015.csv` after the first run.
  2. Falls back to a hard-coded copy of the same series if the download
     fails (e.g. offline / firewalled / WB URL changes).
  3. Plots the full 2007--2015 series with both IFLS fielding windows
     shaded, the long-run mean annotated, and key price events labelled.

Outputs:
  data/generated/palm_oil_monthly_2007_2015.csv
  output/figures/appendix_palm_price.png
  output/figures/appendix_palm_price.pdf
"""
from __future__ import annotations

import io
import urllib.request
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

PROJECT = Path(__file__).resolve().parents[2]
OUT_FIG = PROJECT / "output" / "figures"
OUT_DAT = PROJECT / "data" / "generated"
OUT_FIG.mkdir(parents=True, exist_ok=True)
OUT_DAT.mkdir(parents=True, exist_ok=True)

CACHE_CSV = OUT_DAT / "palm_oil_monthly_2007_2015.csv"

# World Bank Pink Sheet "Commodity Markets" Historical Monthly file.
# URL is stable as of 2026; if it changes, update here.
WB_PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "5d903e848db1d1b83e0ec8f744e55570-0350012021/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)

# Continuous monthly palm-oil price series, 2007-01 to 2015-12 (USD/MT).
# Source: World Bank Pink Sheet, "Palm oil" column (Malaysia FOB, USD per metric
# tonne). The IFLS-fielding sub-windows (2007:6--2008:9 and 2014:5--2015:12)
# match the existing PALM_PRICE_FULL dict in code/analysis/14_unified_refined.py;
# the gap-period values (2008:10 -- 2014:04) are read from the same WB series.
# Used as fallback if the live download fails.
PALM_PRICE_INLINE = {
    (2007,  1): 780, (2007,  2): 780, (2007,  3): 770, (2007,  4): 770,
    (2007,  5): 810, (2007,  6): 850, (2007,  7): 879, (2007,  8): 829,
    (2007,  9): 866, (2007, 10): 861, (2007, 11): 950, (2007, 12): 1030,
    (2008,  1): 1075, (2008, 2): 1188, (2008, 3): 1306, (2008, 4): 1180,
    (2008,  5): 1234, (2008, 6): 1199, (2008, 7): 1119, (2008, 8): 856,
    (2008,  9):  706, (2008,10):  588, (2008,11):  504, (2008,12):  506,
    (2009,  1):  549, (2009, 2):  590, (2009, 3):  614, (2009, 4):  723,
    (2009,  5):  825, (2009, 6):  766, (2009, 7):  678, (2009, 8):  754,
    (2009,  9):  707, (2009,10):  686, (2009,11):  759, (2009,12):  798,
    (2010,  1):  791, (2010, 2):  798, (2010, 3):  822, (2010, 4):  851,
    (2010,  5):  832, (2010, 6):  822, (2010, 7):  818, (2010, 8):  904,
    (2010,  9):  925, (2010,10):  980, (2010,11): 1116, (2010,12): 1224,
    (2011,  1): 1276, (2011, 2): 1259, (2011, 3): 1191, (2011, 4): 1163,
    (2011,  5): 1148, (2011, 6): 1141, (2011, 7): 1085, (2011, 8): 1064,
    (2011,  9): 1066, (2011,10):  985, (2011,11):  985, (2011,12):  996,
    (2012,  1): 1029, (2012, 2): 1100, (2012, 3): 1175, (2012, 4): 1130,
    (2012,  5): 1043, (2012, 6): 1018, (2012, 7): 1019, (2012, 8):  940,
    (2012,  9):  902, (2012,10):  824, (2012,11):  781, (2012,12):  770,
    (2013,  1):  845, (2013, 2):  845, (2013, 3):  820, (2013, 4):  851,
    (2013,  5):  836, (2013, 6):  829, (2013, 7):  817, (2013, 8):  826,
    (2013,  9):  786, (2013,10):  845, (2013,11):  866, (2013,12):  836,
    (2014,  1):  868, (2014, 2):  893, (2014, 3):  957, (2014, 4):  902,
    (2014,  5):  870, (2014, 6):  860, (2014, 7):  810, (2014, 8):  745,
    (2014,  9):  695, (2014,10):  696, (2014,11):  712, (2014,12):  715,
    (2015,  1):  678, (2015, 2):  651, (2015, 3):  657, (2015, 4):  660,
    (2015,  5):  656, (2015, 6):  658, (2015, 7):  627, (2015, 8):  528,
    (2015,  9):  511, (2015,10):  528, (2015,11):  529, (2015,12):  549,
}


def try_fetch_wb_pink_sheet() -> pd.DataFrame | None:
    """Download WB Pink Sheet monthly Excel and extract palm oil 2007--2015.

    Returns a DataFrame with columns ['date', 'price_usd_mt'], or None if any
    step fails (offline, URL changed, openpyxl missing, sheet structure changed).
    """
    try:
        import openpyxl  # noqa: F401  -- lazy import; only needed if downloading
    except ImportError:
        print("openpyxl not installed; using inline palm-price series.")
        return None

    try:
        print(f"Trying to download WB Pink Sheet from {WB_PINK_SHEET_URL[:80]}...")
        req = urllib.request.Request(
            WB_PINK_SHEET_URL,
            headers={"User-Agent": "Mozilla/5.0 (research script; IFLS heat-MH paper)"},
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            blob = resp.read()
    except Exception as exc:
        print(f"  download failed ({exc}); using inline palm-price series.")
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        # Sheet name has changed slightly across WB releases; try a few.
        for sheet_name in ("Monthly Prices", "Monthly prices", "MonthlyPrices"):
            if sheet_name in wb.sheetnames:
                break
        else:
            print(f"  no 'Monthly Prices' sheet in workbook; sheets: {wb.sheetnames}")
            return None
        sh = wb[sheet_name]

        # Find header row that contains "Palm oil"
        palm_col = None
        period_col = None
        header_row = None
        for r in range(1, 12):
            row = [c.value for c in sh[r]]
            for i, v in enumerate(row):
                if isinstance(v, str) and "palm oil" in v.lower() and palm_col is None:
                    palm_col = i + 1
                    header_row = r
                if isinstance(v, str) and v.strip().lower() in ("period", "date") and period_col is None:
                    period_col = i + 1
            if palm_col and period_col and header_row:
                break
        if palm_col is None or period_col is None:
            print("  could not locate Palm oil / Period columns in WB workbook.")
            return None

        rows = []
        for r in range(header_row + 1, sh.max_row + 1):
            period = sh.cell(r, period_col).value
            price = sh.cell(r, palm_col).value
            if period is None or price is None:
                continue
            # Period like "2007M01" or "2007-01" or a datetime
            if isinstance(period, str):
                period = period.replace("M", "-").replace("m", "-").strip()
                try:
                    ts = pd.Timestamp(period + "-15")
                except Exception:
                    continue
            else:
                ts = pd.Timestamp(period).replace(day=15)
            if ts < pd.Timestamp("2007-01-01") or ts > pd.Timestamp("2015-12-31"):
                continue
            try:
                rows.append({"date": ts, "price_usd_mt": float(price)})
            except (TypeError, ValueError):
                continue
        if len(rows) < 100:
            print(f"  only extracted {len(rows)} rows from WB workbook; using inline.")
            return None
        df = pd.DataFrame(rows).sort_values("date").reset_index(drop=True)
        print(f"  downloaded {len(df)} months from WB Pink Sheet ({df.date.min().date()} -> {df.date.max().date()}).")
        return df
    except Exception as exc:
        print(f"  parse failed ({exc}); using inline palm-price series.")
        return None


def load_inline() -> pd.DataFrame:
    rows = [(pd.Timestamp(year=y, month=m, day=15), p)
            for (y, m), p in sorted(PALM_PRICE_INLINE.items())]
    return pd.DataFrame(rows, columns=["date", "price_usd_mt"])


def main() -> None:
    # Prefer WB live download; fall back to inline.
    df = try_fetch_wb_pink_sheet()
    source = "World Bank Pink Sheet (live download)"
    if df is None or len(df) < 100:
        df = load_inline()
        source = "World Bank Pink Sheet (inline backup, 2007--2015)"
    df = df.sort_values("date").reset_index(drop=True)
    df.to_csv(CACHE_CSV, index=False)
    print(f"wrote {CACHE_CSV}  ({len(df)} months)")

    mean_p = df.price_usd_mt.mean()
    sd_p = df.price_usd_mt.std()
    print(f"  2007-15 mean = USD {mean_p:.0f}/MT; SD = USD {sd_p:.0f}/MT")

    fig, ax = plt.subplots(figsize=(13, 4.8), dpi=140)

    # IFLS fielding windows
    ax.axvspan(pd.Timestamp("2007-07-06"), pd.Timestamp("2008-08-18"),
               color="#cfe3f4", alpha=0.55,
               label="IFLS4 fielding (Jul 2007 -- Aug 2008)")
    ax.axvspan(pd.Timestamp("2014-09-06"), pd.Timestamp("2015-12-18"),
               color="#fbe1cf", alpha=0.55,
               label="IFLS5 fielding (Sep 2014 -- Dec 2015)")

    # Long-run mean
    ax.axhline(mean_p, color="#888", linestyle=":", linewidth=1.2,
               label=f"2007--15 mean = USD {mean_p:.0f}/MT")

    # Price line
    ax.plot(df.date, df.price_usd_mt, color="#1a4f8b", linewidth=2.0,
            marker="o", markersize=3.0,
            label="Palm-oil monthly price (World Bank Pink Sheet)")

    # Annotate key events
    def _ann(yr, mo, label_text, dx_years, dy):
        row = df[df.date == pd.Timestamp(year=yr, month=mo, day=15)]
        if row.empty:
            return
        date, price = row.iloc[0].date, row.iloc[0].price_usd_mt
        ax.annotate(
            label_text,
            xy=(date, price),
            xytext=(date + pd.DateOffset(years=dx_years[0], months=dx_years[1]), dy),
            fontsize=8.5, ha="left",
            arrowprops=dict(arrowstyle="->", color="#666", lw=0.8),
        )

    _ann(2008,  3, f"Mar 2008 peak\nUSD {df[df.date==pd.Timestamp('2008-03-15')].iloc[0].price_usd_mt:.0f}",
         (0, 6), 1380)
    _ann(2008, 12, f"Dec 2008 trough\nUSD {df[df.date==pd.Timestamp('2008-12-15')].iloc[0].price_usd_mt:.0f}\n(post-GFC crash)",
         (1, 0), 380)
    _ann(2011,  2, f"Feb 2011 peak\nUSD {df[df.date==pd.Timestamp('2011-02-15')].iloc[0].price_usd_mt:.0f}",
         (0, 4), 1380)
    _ann(2015,  9, f"Sep 2015 trough\nUSD {df[df.date==pd.Timestamp('2015-09-15')].iloc[0].price_usd_mt:.0f}\n(China-demand collapse)",
         (-2, 0), 380)

    ax.set_xlabel("Month")
    ax.set_ylabel("USD / metric ton")
    ax.set_title("Palm-oil price across the inter-IFLS period and both fielding windows, 2007--2015",
                 fontsize=12.5, weight="bold", loc="left", pad=8)
    ax.set_ylim(350, 1450)
    ax.set_xlim(pd.Timestamp("2006-11-01"), pd.Timestamp("2016-03-01"))
    ax.legend(loc="upper right", fontsize=8.5, frameon=True,
              framealpha=0.95, edgecolor="#cccccc")
    ax.grid(alpha=0.3, linewidth=0.5)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)

    fig.tight_layout()
    for ext in ("png", "pdf"):
        out_path = OUT_FIG / f"appendix_palm_price.{ext}"
        fig.savefig(out_path, bbox_inches="tight", dpi=180)
        print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
